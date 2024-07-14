from openpyxl import Workbook, load_workbook
from typing import List

class Excel:
    __filepath: str
    __sheetname: (str | None) = None
    __workbook: Workbook
    __worksheet: (Workbook | None) = None

    def get_filepath(self) -> str:
        return self.__filepath
    
    def set_filepath(self, filepath: str) -> None:
        self.__filepath = filepath
        self.__workbook = load_workbook(filepath)

    def set_sheetname(self, sheetname: str) -> None:
        self.__sheetname = sheetname
        self.__worksheet = self.__workbook[sheetname]

    def get_sheetname(self) -> (str | None):
        return self.__sheetname
    
    def get_sheetnames(self) -> List[str]:
        return self.__workbook.sheetnames
    
    def get_workbook(self) -> Workbook:
        return self.__workbook
    
    def get_worksheet(self) -> (Workbook | None):
        return self.__worksheet

    def save_styles(self) -> None:
        self.__workbook.save(self.__filepath)
    


