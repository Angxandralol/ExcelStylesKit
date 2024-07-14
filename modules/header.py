from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color
from modules.excel import Excel
from utils.alphabet import alphabet_num, alphabet_str

class Header(Excel):
    __header = {
        'start_row': 1, 
        'start_column': 'A', 
        'end_row': 2, 
        'end_column': 'B'
    }

    def __init__(self, filepath: str, sheetname: (str | None)=None) -> None:
        self.set_filepath(filepath)
        if sheetname: self.set_sheetname(sheetname)

    def get_header_values(self) -> dict:
        return self.__header

    def set_header(self, start_row: int, start_column: str, end_row: int, end_column: str) -> None:
        self.__header['start_row'] = start_row
        self.__header['start_column'] = start_column
        self.__header['end_row'] = end_row
        self.__header['end_column'] = end_column

    def style_font(self, name='Time New Roman', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000') -> None:
        font_config = Font(name=name, 
                           size=size, 
                           bold=bold, 
                           italic=italic, 
                           vertAlign=vertAlign, 
                           underline=underline, 
                           strike=strike, 
                           color=color)
        if not self.get_sheetname():
            sheetnames = self.get_sheetnames()
            for sheetname in sheetnames:
                int_start_col = alphabet_str[self.__header['start_column']]
                int_end_col = alphabet_str[self.__header['end_column']]
                for index_col in range(int_start_col, int_end_col):
                    for index_row in range(self.__header['start_row'], self.__header['end_row']):
                        wb = self.get_workbook()                
                        cell = wb[sheetname][f'{alphabet_num[index_col]}{index_row}']
                        cell.font = font_config
        else:
            int_start_col = alphabet_str[self.__header['start_column']]
            int_end_col = alphabet_str[self.__header['end_column']]
            for index_col in range(int_start_col, int_end_col):
                for index_row in range(self.__header['start_row'], self.__header['end_row']):                
                    cell = self.get_worksheet()[f'{alphabet_num[index_col]}{index_row}']
                    cell.font = font_config
        self.save_styles()

    def style_bg_color(self, color='FFFFFF') -> None:
        color = Color(rgb=color)
        fill = PatternFill(patternType='solid', fgColor=color)
        if not self.get_sheetname():
            sheetnames = self.get_sheetnames()
            for sheetname in sheetnames:
                int_start_col = alphabet_str[self.__header['start_column']]
                int_end_col = alphabet_str[self.__header['end_column']]
                for index_col in range(int_start_col, int_end_col):
                    for index_row in range(self.__header['start_row'], self.__header['end_row']):
                        wb = self.get_workbook()                
                        cell = wb[sheetname][f'{alphabet_num[index_col]}{index_row}']
                        cell.fill = fill
        else:
            int_start_col = alphabet_str[self.__header['start_column']]
            int_end_col = alphabet_str[self.__header['end_column']]
            for index_col in range(int_start_col, int_end_col):
                for index_row in range(self.__header['start_row'], self.__header['end_row']):                
                    cell = self.get_worksheet()[f'{alphabet_num[index_col]}{index_row}']
                    cell.fill = fill
        self.save_styles()

    def style_aligment(self, horizontal='left', vertical='center', text_rotation=0, wrap_text=False, shrink=False, indent=0) -> None:
        aligment = Alignment(horizontal=horizontal,
                             vertical=vertical, 
                             text_rotation=text_rotation, 
                             wrap_text=wrap_text, 
                             shrink_to_fit=shrink, 
                             indent=indent)
        if not self.get_sheetname():
            sheetnames = self.get_sheetnames()
            for sheetname in sheetnames:
                int_start_col = alphabet_str[self.__header['start_column']]
                int_end_col = alphabet_str[self.__header['end_column']]
                for index_col in range(int_start_col, int_end_col):
                    for index_row in range(self.__header['start_row'], self.__header['end_row']):
                        wb = self.get_workbook()                
                        cell = wb[sheetname][f'{alphabet_num[index_col]}{index_row}']
                        cell.alignment = aligment
        else:
            int_start_col = alphabet_str[self.__header['start_column']]
            int_end_col = alphabet_str[self.__header['end_column']]
            for index_col in range(int_start_col, int_end_col):
                for index_row in range(self.__header['start_row'], self.__header['end_row']):                
                    cell = self.get_worksheet()[f'{alphabet_num[index_col]}{index_row}']
                    cell.alignment = aligment
        self.save_styles()