from modules.header import Header
from modules.content import Content
from openpyxl.styles import Border, Side
from utils.alphabet import alphabet_num, alphabet_str

class TableConfig:
    __cells: dict
    header: Header
    content: Content

    def __init__(self, filepath: str, sheetname: (str | None)=None) -> None:
        self.header = Header(filepath, sheetname)
        self.content = Content(filepath, sheetname)
        self.__total_rows()

    def __total_rows(self) -> None:
        header = self.header.get_header_values()
        content = self.content.get_content_values()
        start_row = header['start_row']
        start_header_col = alphabet_str[header['start_column']]
        start_content_col = alphabet_str[content['start_column']]
        if start_header_col != start_content_col:
            if start_content_col < start_header_col: start_col = start_content_col
            else: start_col = start_header_col
        else: start_col = start_header_col
        end_row = content['end_row']
        end_header_col = alphabet_str[header['end_column']]
        end_content_col = alphabet_str[content['end_column']]
        if end_header_col != end_content_col:
            if end_content_col < end_header_col: end_col = end_content_col
            else: end_col = end_header_col
        else: end_col = end_header_col
        cells = {
            'start_row': start_row,
            'start_column': alphabet_num[start_col], 
            'end_row': end_row + 1,
            'end_column': alphabet_num[end_col + 1]
        }
        self.__cells = cells

    def get_total_cells(self) -> dict:
        return self.__cells

    def style_border(self, color='000000', style='thin') -> None:
        self.__total_rows()
        border = Border(left=Side(border_style=style, color=color),
                        right=Side(border_style=style, color=color),
                        top=Side(border_style=style, color=color), 
                        bottom=Side(border_style=style, color=color), 
                        diagonal=Side(border_style=style, color=color), 
                        diagonal_direction=0,
                        outline=Side(border_style=style, color=color),
                        vertical=Side(border_style=style, color=color), 
                        horizontal=Side(border_style=style, color=color))
        if not self.content.get_sheetname():
            sheetnames = self.content.get_sheetnames()
            for sheetname in sheetnames:
                int_start_col = alphabet_str[self.__cells['start_column']]
                int_end_col = alphabet_str[self.__cells['end_column']]
                for index_col in range(int_start_col, int_end_col):
                    for index_row in range(self.__cells['start_row'], self.__cells['end_row']):
                        wb = self.content.get_workbook()                
                        cell = wb[sheetname][f'{alphabet_num[index_col]}{index_row}']
                        cell.border = border
        else:
            int_start_col = alphabet_str[self.__cells['start_column']]
            int_end_col = alphabet_str[self.__cells['end_column']]
            for index_col in range(int_start_col, int_end_col):
                for index_row in range(self.__cells['start_row'], self.__cells['end_row']):
                    wb = self.content.get_workbook()                
                    cell = wb[sheetname][f'{alphabet_num[index_col]}{index_row}']
                    cell.border = border
        self.content.save_styles()

    def style_column_width(self, width=30) -> None:
        if not self.content.get_sheetname():
            sheetnames = self.content.get_sheetnames()
            for sheetname in sheetnames:
                sheet = self.content.get_workbook()[sheetname]
                int_start_col = alphabet_str[self.__cells['start_column']]
                int_end_col = alphabet_str[self.__cells['end_column']] + 1 
                for index in range(int_start_col, int_end_col):
                    sheet.column_dimensions[alphabet_num[index]].width = width
        else:
            sheet = self.content.get_workbook()[self.content.get_sheetname()]
            int_start_col = alphabet_str[self.__cells['start_column']]
            int_end_col = alphabet_str[self.__cells['end_column']] + 1
            for index in range(int_start_col, int_end_col):
                sheet.column_dimensions[index].width = width
        self.content.save_styles()